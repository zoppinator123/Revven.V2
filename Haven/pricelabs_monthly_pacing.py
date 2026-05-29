#!/usr/bin/env python3
"""Build monthly pacing tasks from a PriceLabs Report Builder CSV export."""

from __future__ import annotations

import csv
import re
import uuid
from datetime import date, datetime, timezone
from pathlib import Path
from typing import Any


SOURCE = "pricelabs_report_builder_monthly"


def _norm(value: str) -> str:
    return re.sub(r"[^a-z0-9]+", "", str(value or "").lower())


def _property_key(value: str) -> str:
    value = str(value or "").split(" -- ")[0].split(": Default")[0]
    return _norm(value)


def _parse_money(value: Any) -> float | None:
    text = str(value or "").strip()
    if not text or text in {"-", "NA", "N/A"}:
        return None
    text = text.replace("$", "").replace(",", "").replace("(", "-").replace(")", "")
    try:
        return float(text)
    except ValueError:
        return None


def _parse_pct(value: Any) -> float | None:
    text = str(value or "").strip()
    if not text or text in {"-", "NA", "N/A"}:
        return None
    text = text.replace("%", "").replace("+", "").replace(",", "").replace("pts", "").strip()
    try:
        val = float(text)
    except ValueError:
        return None
    return val / 100 if abs(val) > 1 else val


def _pick(row: dict[str, str], aliases: list[str]) -> str:
    normalized = {_norm(k): v for k, v in row.items()}
    for alias in aliases:
        if _norm(alias) in normalized:
            return normalized[_norm(alias)]
    return ""


def _money(value: float | None) -> str:
    return "-" if value is None else f"${value:,.0f}"


def _pct(value: float | None, signed: bool = False) -> str:
    if value is None:
        return "-"
    prefix = "+" if signed and value > 0 else ""
    return f"{prefix}{value * 100:.1f}%"


def _round_to_5(value: float) -> int:
    return max(0, int(round(value / 5) * 5))


def _classify(revenue_yoy: float | None, occ: float | None, occ_stly: float | None, occ_gap: float | None) -> tuple[str, str, str, float | None]:
    gap = occ_gap
    if gap is None and occ is not None and occ_stly is not None:
        gap = occ - occ_stly

    if revenue_yoy is not None and revenue_yoy <= -0.10 and ((gap is not None and gap <= -0.05) or (occ is not None and occ <= 0.35)):
        pct = -0.05 if occ is not None and occ <= 0.20 else -0.03
        return ("behind_both", "high", "Behind revenue and occupancy", pct)
    if (gap is not None and gap <= -0.08) or (occ is not None and occ <= 0.25):
        pct = -0.05 if occ is not None and occ <= 0.20 else -0.03
        return ("behind_occupancy", "high", "Behind occupancy", pct)
    if revenue_yoy is not None and revenue_yoy <= -0.15:
        return ("behind_revenue", "medium", "Behind revenue", None)
    if (revenue_yoy is not None and revenue_yoy >= 0.10 and gap is not None and gap >= 0.05) or (occ is not None and occ >= 0.75):
        pct = 0.05 if occ is not None and occ >= 0.85 else 0.03
        return ("overperforming", "medium", "Over pacing", pct)
    return ("monitor", "low", "Monitor", None)


def _target_month_range(today: date) -> tuple[date, date, str]:
    """Return (start, end, label) for the next calendar month."""
    import calendar
    year, month = (today.year, today.month + 1) if today.month < 12 else (today.year + 1, 1)
    last_day = calendar.monthrange(year, month)[1]
    start = date(year, month, 1)
    end = date(year, month, last_day)
    label = start.strftime("%B %Y")
    return start, end, label


def _action_for_row(row: dict[str, str], prop: Any | None, today: date, report_label: str) -> dict[str, Any]:
    listing = _pick(row, ["Listing Name", "Listing", "Property", "Property Name"]) or (getattr(prop, "name", "") if prop else "Unknown")
    revenue = _parse_money(_pick(row, ["Rental Revenue", "Revenue"]))
    revenue_stly = _parse_money(_pick(row, ["Rental Revenue STLY", "Revenue STLY", "Rental Revenue Same Time Last Year"]))
    revenue_yoy = _parse_pct(_pick(row, ["Rental Revenue STLY YoY %", "Rental Revenue YoY %", "Revenue YoY %", "Rental Revenue YOY"]))
    occ = _parse_pct(_pick(row, ["Paid Occupancy %", "Paid Occupancy", "Occupancy %"]))
    occ_stly = _parse_pct(_pick(row, ["Paid Occupancy STLY", "Paid Occupancy STLY %", "Occupancy STLY"]))
    occ_gap = _parse_pct(_pick(row, ["Paid Occupancy STLY YOY Difference", "Paid Occupancy STLY YoY Difference", "Paid Occupancy YoY Difference", "Occupancy YoY Difference"]))
    revpar = _parse_money(_pick(row, ["Rental RevPAR", "RevPAR"]))
    market_revpar = _parse_money(_pick(row, ["Market RevPAR"]))

    signal, priority, title, pct = _classify(revenue_yoy, occ, occ_stly, occ_gap)
    base = float(getattr(prop, "base_price", 0) or 0)
    min_price = float(getattr(prop, "min_price", 0) or 0)

    month_start, month_end, month_label = _target_month_range(today)
    date_range_str = f"{month_start.strftime('%b %-d')}–{month_end.strftime('%-d, %Y')}"

    hit_floor = False
    proposed_override = None
    if pct is not None and base > 0:
        proposed_override = _round_to_5(base * (1 + pct))
        if pct < 0 and min_price and proposed_override <= min_price < base:
            proposed_override = _round_to_5(min_price)
            hit_floor = True

    if signal in {"behind_both", "behind_occupancy"} and pct is not None:
        suggestion = f"Set monthly override for {month_label}: {pct:+.0%}" if not hit_floor else f"Review minimum floor before setting {month_label} override"
        adjustment = "Minimum floor check" if hit_floor else f"{pct:+.0%} override · {month_label}"
        proposed = f"Date override {pct:+.0%} (est. {_money(proposed_override)}) · {date_range_str}" if proposed_override else f"Date override {pct:+.0%} · {date_range_str}"
        implementation = (
            f"Approve first. In PriceLabs, open this listing's calendar, select all dates from "
            f"{date_range_str}, and apply a {pct:+.0%} date-specific override — do not change the base price. "
            f"This limits the discount to {month_label} only."
        )
        payload_kind = "monthly_date_override" if prop and proposed_override and not hit_floor else "manual_review"
    elif signal == "behind_revenue":
        suggestion = "Investigate revenue gap before applying override"
        adjustment = "No automatic override"
        proposed = "Check ADR, restrictions, channel mix, and open-date exposure"
        implementation = "Revenue is behind, but occupancy is not weak enough for a monthly override."
        payload_kind = "manual_review"
    elif signal == "overperforming" and pct is not None:
        suggestion = f"Set monthly override for {month_label}: {pct:+.0%}"
        adjustment = f"{pct:+.0%} override · {month_label}"
        proposed = f"Date override {pct:+.0%} (est. {_money(proposed_override)}) · {date_range_str}" if proposed_override else f"Date override {pct:+.0%} · {date_range_str}"
        implementation = (
            f"Approve first. In PriceLabs, open this listing's calendar, select all dates from "
            f"{date_range_str}, and apply a {pct:+.0%} date-specific override — do not change the base price. "
            f"Recheck pickup after 7 days."
        )
        payload_kind = "monthly_date_override" if prop and proposed_override else "manual_review"
    else:
        suggestion = "Monitor monthly pacing"
        adjustment = "No override needed"
        proposed = "Keep current setup; watch revenue and paid occupancy next refresh"
        implementation = "No task unless the next report shows a larger revenue or occupancy gap."
        payload_kind = "manual_review"

    reason_bits = [
        f"Revenue {_money(revenue)} vs STLY {_money(revenue_stly)} ({_pct(revenue_yoy, True)} YoY)",
        f"paid occupancy {_pct(occ)} vs STLY {_pct(occ_stly)} ({_pct(occ_gap, True)} gap)",
    ]
    if revpar is not None or market_revpar is not None:
        reason_bits.append(f"RevPAR {_money(revpar)} vs market {_money(market_revpar)}")

    action = {
        "id": f"monthly::{_property_key(listing)}::{_norm(report_label)}::{signal}",
        "property": getattr(prop, "name", listing) if prop else listing,
        "display_name": listing,
        "listing_id": getattr(prop, "listing_id", "") if prop else "",
        "pms_name": getattr(prop, "pms_name", "") if prop else "",
        "group": getattr(prop, "customization_group", "") if prop else "",
        "subgroup": getattr(prop, "customization_sub_group", "") if prop else "",
        "city": getattr(prop, "city", "") if prop else "",
        "group_label": " / ".join(v for v in [getattr(prop, "customization_group", ""), getattr(prop, "customization_sub_group", "")] if v) if prop else "",
        "system": "PriceLabs Report Builder",
        "source": SOURCE,
        "type": f"monthly_{signal}",
        "priority": priority,
        "status": "pending",
        "created_at": datetime.now(timezone.utc).isoformat(),
        "reviewed_at": None,
        "target_dates": f"{month_label} ({date_range_str})",
        "suggestion": suggestion,
        "current_value": f"Base {_money(base) if base else '-'}; revenue YoY {_pct(revenue_yoy, True)}; occ {_pct(occ)}",
        "proposed_value": proposed,
        "adjustment": adjustment,
        "reason": "; ".join(reason_bits),
        "implementation": implementation,
        "monthly_signal": signal,
        "override_month": month_label,
        "override_start": month_start.isoformat(),
        "override_end": month_end.isoformat(),
        "monthly_metrics": {
            "rental_revenue": revenue,
            "rental_revenue_stly": revenue_stly,
            "rental_revenue_yoy": revenue_yoy,
            "paid_occupancy": occ,
            "paid_occupancy_stly": occ_stly,
            "paid_occupancy_gap": occ_gap,
            "revpar": revpar,
            "market_revpar": market_revpar,
        },
        "pricelabs_payload": {
            "kind": payload_kind,
            "adjustment_pct": pct,
            "override_start": month_start.isoformat(),
            "override_end": month_end.isoformat(),
            "estimated_override_price": proposed_override,
        },
    }
    if payload_kind == "manual_review":
        action["pricelabs_payload"] = {"kind": "manual_review"}
    return action


def _iter_rows(path: Path):
    """Yield rows as str-keyed dicts from either a CSV or an XLSX file."""
    with path.open("rb") as fb:
        magic = fb.read(4)
    if magic[:2] == b"PK":
        import io
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(path.read_bytes()), read_only=True, data_only=True)
        ws = wb.active
        row_iter = iter(ws.rows)
        headers = [str(c.value or "").strip() for c in next(row_iter)]
        for row in row_iter:
            yield {headers[i]: str(c.value if c.value is not None else "") for i, c in enumerate(row) if i < len(headers)}
        wb.close()
    else:
        with path.open("r", encoding="utf-8-sig", errors="replace", newline="") as f:
            for row in csv.DictReader(f):
                yield row


def load_monthly_pacing(csv_path: Path, portfolio: list[Any], today: date | None = None) -> dict[str, Any]:
    today = today or date.today()
    if not csv_path.exists():
        return {"ok": False, "error": "No PriceLabs Report Builder CSV uploaded yet.", "actions": [], "summary": {}}

    by_key: dict[str, Any] = {}
    for prop in portfolio:
        if getattr(prop, "active", False):
            by_key[_property_key(getattr(prop, "name", ""))] = prop
            by_key[_property_key(getattr(prop, "property_name", ""))] = prop

    actions: list[dict[str, Any]] = []
    unmatched = 0
    for row in _iter_rows(csv_path):
        listing = _pick(row, ["Listing Name", "Listing", "Property", "Property Name"])
        if not listing:
            continue
        prop = by_key.get(_property_key(listing))
        if not prop:
            unmatched += 1
        actions.append(_action_for_row(row, prop, today, "Report Builder range"))

    rank = {"behind_both": 0, "behind_occupancy": 1, "behind_revenue": 2, "overperforming": 3, "monitor": 4}
    actions.sort(key=lambda a: (rank.get(a.get("monthly_signal"), 9), a.get("property", "")))
    counts: dict[str, int] = {}
    for action in actions:
        counts[action.get("monthly_signal", "monitor")] = counts.get(action.get("monthly_signal", "monitor"), 0) + 1
    return {
        "ok": True,
        "actions": actions,
        "summary": {
            "total": len(actions),
            "unmatched": unmatched,
            "behind_both": counts.get("behind_both", 0),
            "behind_occupancy": counts.get("behind_occupancy", 0),
            "behind_revenue": counts.get("behind_revenue", 0),
            "overperforming": counts.get("overperforming", 0),
            "monitor": counts.get("monitor", 0),
        },
    }
